import os
import sys
import datetime
import openpyxl
from dateutil.parser import parse
from subprocess import call
sdate = "04/01/2017"

def add_data(no_days, no_rows):
    for j in range(1, 97):
        owb = openpyxl.load_workbook("Market_Snapshot_96/newbidFile"+str(j)+".xlsx")
        osheet = owb.get_sheet_by_name('Sheet')
    	wwb = openpyxl.Workbook()
    	wsheet = wwb.get_sheet_by_name('Sheet')
    	print "opened newbidFile"+str(j)+".xlsx"
        for i in range(1, no_rows + 1):
            wsheet['A'+str(i)] = osheet['A'+str(i)].value
            wsheet['B'+str(i)] = osheet['B'+str(i)].value
            wsheet['C'+str(i)] = osheet['C'+str(i)].value
            wsheet['D'+str(i)] = osheet['D'+str(i)].value
            wsheet['E'+str(i)] = osheet['E'+str(i)].value
            wsheet['F'+str(i)] = osheet['F'+str(i)].value
            wsheet['G'+str(i)] = osheet['G'+str(i)].value
        owb = openpyxl.load_workbook("BidFile1.xlsx")
        osheet = owb.get_sheet_by_name('MarketMinute')
        k = 6 + j
        for rc in range(no_rows + 1, no_rows + no_days + 1):
            wsheet['A'+str(rc)] = (parse(sdate)+datetime.timedelta(days=rc-2)).strftime('%d/%m/%Y')
            wsheet['B'+str(rc)] = osheet['D'+str(k)].value
            wsheet['C'+str(rc)] = osheet['F'+str(k)].value
            wsheet['D'+str(rc)] = osheet['G'+str(k)].value
            wsheet['E'+str(rc)] = osheet['H'+str(k)].value
            wsheet['F'+str(rc)] = osheet['J'+str(k)].value
            wsheet['G'+str(rc)] = osheet['K'+str(k)].value
            k  = k + 96
    	print "saving newbidFile"+str(j)+".xlsx......."
    	wwb.save('newbidFile'+str(j)+'.xlsx')

add_data(27, 1896)
